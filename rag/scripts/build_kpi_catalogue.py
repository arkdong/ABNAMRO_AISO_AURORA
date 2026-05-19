"""Export the Content KPI Inventory xlsx into a runtime-friendly JSON.

Reads ``data/Content KPI inventory_AISO.xlsx`` and writes
``backend/evaluation/data/kpi_catalogue.json``. The evaluation module loads
the JSON at startup, so openpyxl is a build-time-only dependency.

Schema (output):

    {
      "source_xlsx": <relative path>,
      "generated_at": <ISO-8601 UTC>,
      "categories": [
        {"name": "...", "weight": "High"}, ...
      ],
      "clusters": [
        {"name": "...", "weight": "Medium", "primary": "..."}, ...
      ],
      "kpis": [
        {
          "id": "factuality_truthfullness",
          "name": "Factuality & truthfullness",
          "primary_cluster": "Compliancy  & substantive quality / accuracy, ...",
          "secondary_cluster": null,
          "category": "Compliancy & substantive quality",
          "cluster_short": "Accuracy, efficacy and reliability",
          "monitoring": "Mandatory",
          "weight": "Blocking",
          "contribution": "...",
          "norm": "...",
          "norm_chat": null,
          "norm_messages": null,
          "measurement": "...",
          "indicator": "ErrorScale",
          "automated_match": "no match (manual check only)",
          "relevance": {
            "human": "Applicable",
            "genai_knowledge": "Applicable",
            "instant": "Applicable",
            "web": "Applicable", "chat": "Applicable", "messages": "Applicable",
            "employee": "Applicable", "app_ib": "Applicable"
          },
          "guardrail_category": "[Missing category]",
          "guardrail_tag": "[meerdere, zie comments]"
        }, ...
      ]
    }

Run:  ``python rag/scripts/build_kpi_catalogue.py``
"""
from __future__ import annotations

import json
import re
import warnings
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = REPO_ROOT / "data" / "Content KPI inventory_AISO.xlsx"
OUT_PATH = REPO_ROOT / "backend" / "evaluation" / "data" / "kpi_catalogue.json"

# Inventory sheet column map (1-indexed, matches the header row at row 8).
COL = {
    "number": 1, "type": 2, "primary_cluster": 3, "secondary_cluster": 4,
    "category_name": 5, "cluster_name": 6, "final_name": 7,
    "monitoring": 8, "weight": 9, "contribution": 10,
    "norm": 11, "norm_source": 12, "norm_chat": 13, "norm_chat_src": 14,
    "norm_messages": 15, "norm_messages_src": 16,
    "measurement": 17, "indicator": 18, "automated_match": 19,
    "rel_human": 20, "rel_genai_knowledge": 21, "rel_instant": 22,
    "rel_web": 23, "rel_chat": 24, "rel_messages": 25,
    "rel_employee": 26, "rel_app_ib": 27,
    "guardrail_category": 28, "guardrail_category_comment": 29,
    "guardrail_tag": 30, "guardrail_tag_comment": 31,
}

# Map free-text indicator strings to the indicator enum names in
# ``backend.evaluation.indicators``. Keep this in lockstep with that module.
INDICATOR_ENUM_BY_PHRASE: dict[str, str] = {
    "low, medium or high maturity": "Maturity",
    "present, not present": "PresenceScale",
    "yes/no deviation from norm": "DeviationYesNo",
    "completed step yes/no": "YesNoScale",
    "yes/no": "YesNoScale",
    "used, not used": "UsedScale",
    "applicable, not applicable": "ApplicableScale",
    "many deviations, few deviations, no deviations": "DeviationScale",
    "many ambiguities, few ambiguities, no ambiguities": "AmbiguityScale",
    "off-topic, somewhat relevant, reasonable relevant, highly relevant": "RelevanceScale",
    "no grounding, limited grounding, reasonable grounding, fully grounded": "GroundednessScale",
    "numerous errors, several errors, moderate errors, few errors, no errors": "ErrorScale",
    "very incomplete, incomplete, fairly complete, mostly complete, fully complete": "CompletenessScale",
    "unclear, somewhat clear, clear, very clear": "ClarityScale",
    "no fit, limited fit, optimal fit": "FitScale",
    "no options, limited options, many options": "OptionsScale",
    "very low, low, medium, high, very high": "FivePointScale",
    "right length, too long": "LengthScale",
    "A1, A2, B1, B2, C1, C2": "LanguageLevelScale",
    "exclusion. no exclusion": "ExclusionScale",
    "published, not published": "PublishedScale",
    "neutral, positive, negative": "SentimentScale",
    "neutral, male. female": "GenderScale",
    "1 very difficult, 2 difficult, 3 neutral, 4 easy, 5 very easy": "CESScale",
    "1 very dissatisfied, 2 dissatisfied, 3 neutral, 4 satisfied, 5 very satisfied related to norm": "CSATScale",
    "0 being not likely at all - 10 very likely to refer": "NPSScale",
}


def _norm_text(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _slugify(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s or "kpi"


def _map_indicator(phrase: str | None) -> str | None:
    if not phrase:
        return None
    key = phrase.strip().lower()
    for k, v in INDICATOR_ENUM_BY_PHRASE.items():
        if k.lower() == key:
            return v
    return None  # unknown indicator → leave null; consumers know to skip the gate


def _derive_category(primary_cluster: str | None) -> str | None:
    if not primary_cluster:
        return None
    # primary cluster strings are formatted "Category / sub-cluster".
    head = primary_cluster.split("/", 1)[0].strip()
    return head or None


def _short_cluster(primary_cluster: str | None) -> str | None:
    if not primary_cluster or "/" not in primary_cluster:
        return primary_cluster
    return primary_cluster.split("/", 1)[1].strip()


def build() -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Inventory"]

    categories: list[dict] = []
    clusters: list[dict] = []
    kpis: list[dict] = []
    used_ids: set[str] = set()

    for ri in range(9, ws.max_row + 1):
        kpi_type = _norm_text(ws.cell(ri, COL["type"]).value)
        if not kpi_type:
            continue
        name = _norm_text(ws.cell(ri, COL["final_name"]).value)
        if not name:
            continue
        weight = _norm_text(ws.cell(ri, COL["weight"]).value) or "Medium"
        # Typo in source: "Mediun" appears once. Normalise.
        if weight.lower() == "mediun":
            weight = "Medium"
        primary_cluster = _norm_text(ws.cell(ri, COL["primary_cluster"]).value)

        if kpi_type == "KPI category":
            categories.append({"name": name, "weight": weight})
            continue
        if kpi_type == "KPI cluster":
            clusters.append(
                {
                    "name": name,
                    "weight": weight,
                    "primary_cluster": primary_cluster,
                }
            )
            continue
        if kpi_type != "KPI":
            continue

        # de-duplicate slugs (a few rows share final_name, e.g. age discrimination).
        base_id = _slugify(name)
        slug = base_id
        n = 2
        while slug in used_ids:
            slug = f"{base_id}_{n}"
            n += 1
        used_ids.add(slug)

        indicator_phrase = _norm_text(ws.cell(ri, COL["indicator"]).value)
        kpi = {
            "id": slug,
            "name": name,
            "primary_cluster": primary_cluster,
            "secondary_cluster": _norm_text(
                ws.cell(ri, COL["secondary_cluster"]).value
            ),
            "category": _derive_category(primary_cluster),
            "cluster_short": _short_cluster(primary_cluster),
            "monitoring": _norm_text(ws.cell(ri, COL["monitoring"]).value) or "Optional",
            "weight": weight,
            "contribution": _norm_text(ws.cell(ri, COL["contribution"]).value),
            "norm": _norm_text(ws.cell(ri, COL["norm"]).value),
            "norm_chat": _norm_text(ws.cell(ri, COL["norm_chat"]).value),
            "norm_messages": _norm_text(ws.cell(ri, COL["norm_messages"]).value),
            "measurement": _norm_text(ws.cell(ri, COL["measurement"]).value),
            "indicator_phrase": indicator_phrase,
            "indicator": _map_indicator(indicator_phrase),
            "automated_match": _norm_text(ws.cell(ri, COL["automated_match"]).value),
            "relevance": {
                "human": _norm_text(ws.cell(ri, COL["rel_human"]).value),
                "genai_knowledge": _norm_text(ws.cell(ri, COL["rel_genai_knowledge"]).value),
                "instant": _norm_text(ws.cell(ri, COL["rel_instant"]).value),
                "web": _norm_text(ws.cell(ri, COL["rel_web"]).value),
                "chat": _norm_text(ws.cell(ri, COL["rel_chat"]).value),
                "messages": _norm_text(ws.cell(ri, COL["rel_messages"]).value),
                "employee": _norm_text(ws.cell(ri, COL["rel_employee"]).value),
                "app_ib": _norm_text(ws.cell(ri, COL["rel_app_ib"]).value),
            },
            "guardrail_category": _norm_text(ws.cell(ri, COL["guardrail_category"]).value),
            "guardrail_tag": _norm_text(ws.cell(ri, COL["guardrail_tag"]).value),
        }
        kpis.append(kpi)

    # Synthetic KPIs for the standalone "Search Quality Raters for GenAI"
    # block (Criteria sheet, cols T/U) which doesn't have its own rows in the
    # Inventory sheet. We materialise them so the Tier 2 GenAI rater judges
    # have catalogue entries to bind to.
    synthetic_genai_rater = [
        {
            "id": "no_paraphrase",
            "name": "No paraphrased content",
            "source": "Criteria sheet — Search quality raters for GenAI",
        },
        {
            "id": "no_filler",
            "name": "No filler / too generic information",
            "source": "Criteria sheet — Search quality raters for GenAI",
        },
    ]
    for s in synthetic_genai_rater:
        if s["id"] in used_ids:
            continue
        used_ids.add(s["id"])
        kpis.append(
            {
                "id": s["id"],
                "name": s["name"],
                "primary_cluster": "Online findability & visibility / content value",
                "secondary_cluster": None,
                "category": "Online findability and visibility",
                "cluster_short": "content value",
                "monitoring": "Optional",
                "weight": "High",
                "contribution": "Easily findable, GenAI-rater-style quality: no paraphrase, no filler.",
                "norm": "obligatory for all generated content",
                "norm_chat": None,
                "norm_messages": None,
                "measurement": "LLM-as-judge check on generated content against rubric.",
                "indicator_phrase": "present, not present",
                "indicator": "PresenceScale",
                "automated_match": "no match (manual check only)",
                "relevance": {
                    "human": "Applicable",
                    "genai_knowledge": "Applicable",
                    "instant": "Applicable",
                    "web": "Applicable",
                    "chat": "Applicable",
                    "messages": "Applicable",
                    "employee": "Applicable",
                    "app_ib": "Applicable",
                },
                "guardrail_category": "[Missing category]",
                "guardrail_tag": None,
                "source_block": s["source"],
            }
        )

    return {
        "source_xlsx": str(XLSX_PATH.relative_to(REPO_ROOT)),
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "categories": categories,
        "clusters": clusters,
        "kpis": kpis,
    }


def main() -> None:
    out = build()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    # Quick visibility on coverage.
    unknown = sum(1 for k in out["kpis"] if k["indicator"] is None)
    blocking = sum(1 for k in out["kpis"] if k["weight"] == "Blocking")
    print(
        f"Wrote {OUT_PATH.relative_to(REPO_ROOT)}: "
        f"{len(out['kpis'])} KPIs, {len(out['categories'])} categories, "
        f"{len(out['clusters'])} clusters, "
        f"{blocking} Blocking, {unknown} with unmapped indicator phrase."
    )


if __name__ == "__main__":
    main()
